"""Importers from the legacy master files into a ThermoDB.

Requires the ``import`` extra (``pandas``, ``openpyxl``) -- the core
``db``/``formation`` modules do not need either.

Three legacy sources exist today, all sharing the same 12-column
experimental-reference block (by *position*; header text drifts slightly
between them -- "ΔfH°(0)" vs "ΔfH°gas(0)" vs "DfH0(0)" for the same column):

- Paul's master workbook, "Atom Reference Energies and States.xlsx" --
  the clean experimental block only, one row per element.
- "VASP element_energies.xlsx" (sheet "element_energies") -- the same
  experimental block, plus paired ``"<method>@<encut> atom energy"``
  (isolated atom, ref_type="atom") and ``"<method>@<encut>"`` (per-atom
  standard-state energy, ref_type="element_phase") columns.
- Each of gaussian_step / psi4_step's ``data/atom_energies.csv`` -- the
  same experimental block, plus one column per method (isolated atom,
  kJ/mol) and an optional ``"<method> correction"`` column.

A fourth, ongoing source needs no xlsx/openpyxl at all:

- ``import_orca_atom_results`` reads a per-job ORCA atom-energy results
  CSV (one row per element, one ``E DFT@<method>/<basis> (kJ/mol)`` /
  ``S^2 DFT@<method>/<basis>`` column pair per method+basis run) and
  vets each cell against its expected spin before adding it -- see that
  function's docstring.
"""

import csv
import math
import re

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

__all__ = [
    "STANDARD_STATE",
    "import_reference_xlsx",
    "import_vasp_workbook",
    "import_wide_method_csv",
    "import_orca_atom_results",
]

# Transcribed from the `standard_state` dict duplicated in gaussian_step,
# psi4_step, and vasp_step (all three are byte-for-byte identical). Only
# covers Z=1-36 (H-Kr), matching today's experimental DfH coverage.
#
# Note: the original dict's entry for Kr is the bare string "(g)" (missing
# the "Kr" symbol) -- an upstream copy-paste bug. Corrected here to
# "Kr(g)"; worth fixing upstream too once this package replaces the copies.
STANDARD_STATE = {
    "H": "1/2 H2(g)",
    "He": "He(g)",
    "Li": "Li(s)",
    "Be": "Be(s)",
    "B": "B(s)",
    "C": "C(s,gr)",
    "N": "1/2 N2(g)",
    "O": "1/2 O2(g)",
    "F": "1/2 F2(g)",
    "Ne": "Ne(g)",
    "Na": "Na(s)",
    "Mg": "Mg(s)",
    "Al": "Al(s)",
    "Si": "Si(s)",
    "P": "P(s)",
    "S": "S(s)",
    "Cl": "1/2 Cl2(g)",
    "Ar": "Ar(g)",
    "K": "K(s)",
    "Ca": "Ca(s)",
    "Sc": "Sc(s)",
    "Ti": "Ti(s)",
    "V": "V(s)",
    "Cr": "Cr(s)",
    "Mn": "Mn(s)",
    "Fe": "Fe(s)",
    "Co": "Co(s)",
    "Ni": "Ni(s)",
    "Cu": "Cu(s)",
    "Zn": "Zn(s)",
    "Ga": "Ga(s)",
    "Ge": "Ge(s)",
    "As": "As(s)",
    "Se": "Se(s)",
    "Br": "1/2 Br2(l)",
    "Kr": "Kr(g)",
}

# Column layout shared (by position) across all three legacy sources.
_EXPERIMENTAL_COLUMNS = (
    "atomic_number",  # 0
    "symbol",  # 1
    "multiplicity",  # 2
    "term_symbol",  # 3
    "dfH0_0K",  # 4
    "dfH0_298K",  # 5
    "dfH0_298K_stderr",  # 6
    "h298_minus_h0_atom",  # 7
    "h298_minus_h0_std_state",  # 8
    "s298_gas",  # 9
    "s298_gas_stderr",  # 10
    "reference",  # 11
)


def _require_openpyxl():
    if openpyxl is None:
        raise ImportError(
            "importers.py needs openpyxl; install the 'import' extra "
            "(pip install seamm_thermochemistry[import])"
        )


def _add_element_row(db, row, *, reference_note=None):
    """row: a tuple/list positionally matching _EXPERIMENTAL_COLUMNS."""
    values = dict(zip(_EXPERIMENTAL_COLUMNS, row))
    symbol = values["symbol"]
    if symbol is None:
        return None
    db.add_element(
        int(values["atomic_number"]),
        symbol,
        multiplicity=values["multiplicity"],
        term_symbol=values["term_symbol"],
        standard_state=STANDARD_STATE.get(symbol),
        dfH0_0K=values["dfH0_0K"],
        dfH0_298K=values["dfH0_298K"],
        dfH0_298K_stderr=values["dfH0_298K_stderr"],
        h298_minus_h0_atom=values["h298_minus_h0_atom"],
        h298_minus_h0_std_state=values["h298_minus_h0_std_state"],
        s298_gas=values["s298_gas"],
        s298_gas_stderr=values["s298_gas_stderr"],
        reference=values["reference"],
        reference_note=reference_note,
    )
    return symbol


def import_reference_xlsx(db, path, sheet="Reference Data", reference_note="JANAF"):
    """Load the experimental block from Paul's master workbook.

    Returns the list of element symbols imported.
    """
    _require_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        imported = []
        for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
            symbol = _add_element_row(db, row, reference_note=reference_note)
            if symbol is not None:
                imported.append(symbol)
        return imported
    finally:
        wb.close()


_ATOM_ENERGY_RE = re.compile(r"^(.*)@(\d+) atom energy$")
_ELEMENT_PHASE_RE = re.compile(r"^(.*)@(\d+)$")


def import_vasp_workbook(
    db, path, sheet="element_energies", reference_note="JANAF", import_elements=True
):
    """Load the VASP element/atom energy workbook.

    Imports the shared experimental block (optional, `import_elements`) plus
    every ``"<method>@<encut> atom energy"`` column as `ref_type="atom"` and
    every matching ``"<method>@<encut>"`` column as `ref_type="element_phase"`,
    with ``settings=f"encut={encut}eV"``. Both column families are already
    in kJ/mol in this workbook (checked against its own "Testing" sheet,
    which carries the raw VASP eV values alongside their kJ/mol conversion)
    -- not eV, despite the encut being in eV.

    Returns a dict {"elements": [...], "atom_energy": n, "element_phase": n}.
    """
    _require_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        elements = []
        n_atom, n_phase = 0, 0
        for row in rows[1:]:
            if row[1] is None:  # blank trailing row
                continue
            if import_elements:
                symbol = _add_element_row(db, row[:12], reference_note=reference_note)
            else:
                symbol = row[1]
            if symbol is None:
                continue
            elements.append(symbol)

            for col_idx, col_name in enumerate(header):
                if not col_name or col_idx >= len(row):
                    continue
                value = row[col_idx]
                if value is None:
                    continue

                m = _ATOM_ENERGY_RE.match(col_name)
                if m:
                    method, encut = m.group(1), m.group(2)
                    db.add_atom_energy(
                        symbol,
                        "vasp",
                        method,
                        value,
                        ref_type="atom",
                        settings=f"encut={encut}eV",
                        units="kJ/mol",
                        source=str(path),
                    )
                    n_atom += 1
                    continue

                m = _ELEMENT_PHASE_RE.match(col_name)
                if m:
                    method, encut = m.group(1), m.group(2)
                    db.add_atom_energy(
                        symbol,
                        "vasp",
                        method,
                        value,
                        ref_type="element_phase",
                        settings=f"encut={encut}eV",
                        units="kJ/mol",
                        source=str(path),
                    )
                    n_phase += 1

        return {"elements": elements, "atom_energy": n_atom, "element_phase": n_phase}
    finally:
        wb.close()


def import_wide_method_csv(
    db,
    path,
    code,
    *,
    methods=None,
    ref_type="atom",
    reference_note=None,
    import_elements=True,
):
    """Load one of the wide per-code CSVs (gaussian_step/psi4_step style).

    Parameters
    ----------
    code : str
        "gaussian" or "psi4" (or whatever this file belongs to).
    methods : list of str, optional
        Restrict to these method columns (by exact header text). If None,
        import every non-empty method column found -- can be thousands of
        columns for the full Gaussian composite-method grid; pass an
        explicit list for anything but a one-off full import.

    Returns a dict {"elements": [...], "methods": [...], "n_energies": n}.
    """
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        rows = list(reader)

    method_columns = [h for h in header[13:] if h and not h.endswith(" correction")]
    if methods is not None:
        wanted = set(methods)
        method_columns = [h for h in method_columns if h in wanted]
    col_index = {name: i for i, name in enumerate(header)}

    elements = []
    n_energies = 0
    # db.batch(): this can be hundreds of thousands of add_atom_energy calls
    # for the full composite-method grid -- one commit (fsync) per call
    # would make that prohibitively slow.
    with db.batch():
        for row in rows:
            if len(row) < 12 or not row[1]:
                continue
            symbol = None
            if import_elements:
                symbol = _add_element_row(db, row[:12], reference_note=reference_note)
            else:
                symbol = row[1]
            if symbol is None:
                continue
            elements.append(symbol)

            for method in method_columns:
                i = col_index[method]
                if i >= len(row) or not row[i]:
                    continue
                energy = float(row[i])
                correction = None
                corr_name = method + " correction"
                if corr_name in col_index:
                    j = col_index[corr_name]
                    if j < len(row) and row[j]:
                        correction = float(row[j])
                db.add_atom_energy(
                    symbol,
                    code,
                    method,
                    energy,
                    ref_type=ref_type,
                    units="kJ/mol",
                    correction=correction,
                    source=str(path),
                )
                n_energies += 1

    return {"elements": elements, "methods": method_columns, "n_energies": n_energies}


_ORCA_ENERGY_COL_RE = re.compile(r"^E DFT@(?P<method>.+)/(?P<basis>.+) \(kJ/mol\)$")


def _expected_s2(multiplicity):
    """S(S+1) for the given multiplicity 2S+1."""
    return (multiplicity**2 - 1) / 4.0


def _classify_spin(s2_value, multiplicity, warn_threshold, reject_threshold):
    """Classify a computed S^2 against the multiplicity's expected value.

    Returns (verdict, detail) where verdict is "ok" | "warn" | "reject" and
    detail is a human-readable reason (None for "ok").
    """
    expected = _expected_s2(multiplicity)

    if s2_value is None:
        if multiplicity == 1:
            return "ok", None  # trivially 0; commonly left unreported
        return "reject", "no S^2 reported for an open-shell atom"

    if expected == 0:
        # Singlet: expect ~0, not a meaningful relative deviation to compute.
        if abs(s2_value) > 0.05:
            return "reject", f"S^2={s2_value:.4f} but multiplicity=1 (expect ~0)"
        return "ok", None

    rel_dev = abs(s2_value - expected) / expected
    detail = f"S^2={s2_value:.4f} vs expected {expected:.4f} ({rel_dev:.0%} off)"
    if rel_dev > reject_threshold:
        return "reject", detail
    if rel_dev > warn_threshold:
        return "warn", detail
    return "ok", None


def import_orca_atom_results(
    db,
    csv_path,
    *,
    code="orca",
    ref_type="atom",
    warn_threshold=0.02,
    reject_threshold=0.20,
    energy_rel_tol=1e-6,
    energy_abs_tol=0.01,
    force=False,
    dry_run=False,
):
    """Import ORCA atom-energy results from a per-job results CSV.

    Expects the shape produced by the atom-energy scan flowchart: one row
    per element (``Atomic Number``, ``Element``, ``Multiplicity``), and one
    ``"E DFT@<method>/<basis> (kJ/mol)"`` / ``"S^2 DFT@<method>/<basis>"``
    column pair per method+basis combination that has been run so far
    (blank cells are jobs not finished/reached yet, not errors). ``method``
    and ``basis`` are stored separately -- ``method`` as the ThermoDB
    ``method``, ``basis`` as ``settings`` -- since ORCA (unlike Gaussian's
    composite-method columns) always keeps them as independent axes.

    Vetting, since a bad SCF root (wrong occupation, not just unconverged)
    is the actual failure mode seen for hard atoms (lanthanides etc.), not
    just a missing value:

    - A computed S^2 is compared to the multiplicity's expected S(S+1).
      Within `warn_threshold` (relative) it's added silently; beyond that
      but within `reject_threshold` it's added but flagged in
      `"spin_warnings"` for manual review; beyond `reject_threshold` it's
      not added at all (`"rejected"`).
    - A missing S^2 is fine for a singlet (trivially 0, commonly not
      reported) but rejected for any open-shell multiplicity -- no way to
      vet it.
    - If the CSV's own multiplicity for an element disagrees with what's
      already in the `element` table (when known), that's noted in
      `"multiplicity_mismatches"` but does not block the import -- it may
      just mean the reference table's own value needs a look.

    Duplicate handling, since re-running elements/methods is expected
    across "hundreds of these": an existing (element, code, method,
    ref_type, settings) value that's already in the database is left
    alone if the new value is close (`"unchanged"`); if it differs by more
    than `energy_rel_tol`/`energy_abs_tol` it is reported in
    `"conflicts"` and, by default, *not* overwritten -- pass `force=True`
    to overwrite anyway (each such case is then also listed in
    `"updated"`).

    Parameters
    ----------
    db : ThermoDB
    csv_path : str or Path
    code : str
        Defaults to "orca"; parameterized in case this CSV shape is ever
        reused for another code.
    ref_type : str
        Defaults to "atom" (isolated gas-phase atom energies -- the only
        thing this scan computes).
    warn_threshold, reject_threshold : float
        Relative S^2 deviation thresholds (see above). Tune per how
        strict you want the gate; there's no universally "right" value,
        the deviations here are your own to judge (e.g. an ~20%-high S^2
        on some lanthanides is real but was still accepted on review in
        this project).
    energy_rel_tol, energy_abs_tol : float
        `math.isclose` tolerances for treating a duplicate energy value
        as "the same result" rather than a conflict.
    force : bool
        Overwrite conflicting existing values instead of just reporting
        them.
    dry_run : bool
        Classify and report everything, but never call
        `db.add_atom_energy` -- preview a CSV before committing it.

    Returns
    -------
    dict with keys "added", "updated", "unchanged", "conflicts",
    "rejected", "spin_warnings", "multiplicity_mismatches" -- each a list
    of small dicts describing what happened, for the caller to report.
    """
    csv_path = str(csv_path)
    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    energy_columns = []  # (energy_col, s2_col, method, basis)
    for col in fieldnames:
        m = _ORCA_ENERGY_COL_RE.match(col)
        if m:
            method, basis = m.group("method"), m.group("basis")
            energy_columns.append((col, f"S^2 DFT@{method}/{basis}", method, basis))

    summary = {
        "added": [],
        "updated": [],
        "unchanged": [],
        "conflicts": [],
        "rejected": [],
        "spin_warnings": [],
        "multiplicity_mismatches": [],
    }

    with db.batch():
        for row in rows:
            symbol = (row.get("Element") or "").strip()
            if not symbol:
                continue
            try:
                multiplicity = int(row["Multiplicity"])
            except (KeyError, ValueError, TypeError):
                summary["rejected"].append(
                    {"symbol": symbol, "reason": "cannot parse Multiplicity"}
                )
                continue

            element_row = db.get_element(symbol=symbol)
            if element_row is not None and element_row["multiplicity"] is not None:
                if int(element_row["multiplicity"]) != multiplicity:
                    summary["multiplicity_mismatches"].append(
                        {
                            "symbol": symbol,
                            "csv_multiplicity": multiplicity,
                            "db_multiplicity": int(element_row["multiplicity"]),
                        }
                    )

            for energy_col, s2_col, method, basis in energy_columns:
                energy_str = (row.get(energy_col) or "").strip()
                if not energy_str:
                    continue  # job not run/reached yet -- not an error

                energy = float(energy_str)
                s2_str = (row.get(s2_col) or "").strip()
                s2_value = float(s2_str) if s2_str else None

                verdict, detail = _classify_spin(
                    s2_value, multiplicity, warn_threshold, reject_threshold
                )
                entry = {
                    "symbol": symbol,
                    "method": method,
                    "basis": basis,
                    "energy": energy,
                    "s2": s2_value,
                }
                if verdict == "reject":
                    summary["rejected"].append({**entry, "reason": detail})
                    continue
                if verdict == "warn":
                    summary["spin_warnings"].append({**entry, "reason": detail})

                existing = db.get_atom_energy(
                    symbol, code, method, ref_type=ref_type, settings=basis
                )
                if existing is None:
                    if not dry_run:
                        db.add_atom_energy(
                            symbol,
                            code,
                            method,
                            energy,
                            ref_type=ref_type,
                            settings=basis,
                            units="kJ/mol",
                            source=csv_path,
                        )
                    summary["added"].append(entry)
                elif math.isclose(
                    existing, energy, rel_tol=energy_rel_tol, abs_tol=energy_abs_tol
                ):
                    summary["unchanged"].append({**entry, "existing": existing})
                else:
                    conflict = {
                        **entry,
                        "existing": existing,
                        "delta": energy - existing,
                    }
                    if force:
                        if not dry_run:
                            db.add_atom_energy(
                                symbol,
                                code,
                                method,
                                energy,
                                ref_type=ref_type,
                                settings=basis,
                                units="kJ/mol",
                                source=csv_path,
                            )
                        summary["updated"].append(conflict)
                    else:
                        summary["conflicts"].append(conflict)

    return summary
